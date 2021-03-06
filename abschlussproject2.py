from selenium import webdriver
import pandas as pd
from bs4 import BeautifulSoup
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.chrome.options import Options



driver = webdriver.Chrome("C:/selenium beispiel/chromedriver.exe")

url = "https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-kliniken-herzogin-elisabeth-braunschweig"
driver.get(url)

content = driver.page_source
beautisoup = BeautifulSoup(content, 'html.parser')

names = []
ratings = []
reviews = []
departments = []
dates = []
titles = []


for ele in beautisoup.find_all("article", attrs={"class": "bewertung"}):
    title = ele.find("h2")
    titles.append(title.text)
    date = ele.find("time")
    dates.append(date.text)
    department = ele.find("a", href=True)
    departments.append(department.text)
    review = ele.find("p")
    reviews.append(review.text)

    rating = ele.find("dd")
    ratings.append(rating.text)

x = len(titles)

# for ele in beautisoup.find_all("section", attrs={"class":"rating"}):


for ele in beautisoup.find_all("div", attrs={"class": "klinik-normal"}):
    name = ele.find("h1")
    names.append(name.text)
names = names * x
#ratings = ratings * x
df = pd.DataFrame({'Klinik_Name': names, 'Title': titles, 'Date': dates,
                   'Department': departments, 'rating': ratings, 'Erfahrungsbericht & Review': reviews})
df.to_csv('D:/refugeeks_project/selenium_py/p1_py/klinik1_liste1.csv',
          index=False, encoding='utf-8')


"""
Link Klinikbewertungen
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-kliniken-herzogin-elisabeth-braunschweig
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-klinik-am-zuckerberg-braunschweig
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-stadtkrankenhaus-wolfsburg
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-krankenhaus-peine
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-krankenhaus-st-martini-duderstadt
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-krankenhaus-henriettenstiftung-hannover
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-eilenriede-klinik-hannover
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-sophien-klinik-hannover
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-agnes-karl-krankenhaus-laatzen
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-klinikum-wahrendorff-sehnde
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-landeskrankenhaus-hildesheim
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-krankenhaus-nienburg
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-stadtkrankenhaus-cuxhaven
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-drk-krankenhaus-seepark-debstedt-langen
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-krankenhaus-buchholz
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-krankenhaus-winsen
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-landeskrankenhaus-lueneburg
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-diakoniekrankenhaus-rotenburg
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-krankenhaus-osterholz-scharmbeck
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-klinik-fallingbostel
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-rehazentrum-soltau
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-krankenhaus-walsrode
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-krankenhaus-buxtehude
https://www.klinikbewertungen.de/klinik-forum/erfahrung-mit-diana-klinik-physikalische-medizin-bad-bevensen

"""
"""
from os import link
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pandas as pd
import requests
from bs4 import BeautifulSoup
import csv
from itertools import zip_longest

import openpyxl
from openpyxl import load_workbook
wb = openpyxl.load_workbook('Klinikliste.xlsx')
sheet = wb['Tabelle1']
url = [] # ist eine liste von allen Kliniken Links
for row in sheet.iter_rows(
     min_row= 2,
     max_row=16,
     min_col=3,
     values_only=True):

     url.append(row)
     
klinik_name = []
bewerbungs_text = []
stern_bewerbung = []
datum_bew = []
klinik_link = []




"""
